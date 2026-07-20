"""Build the user-facing nine-sheet GST reconciliation workbook."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from reconciliation import MatchPair, ReconciliationReport


SHEET_NAMES = [
    "1 Purchase Register",
    "2 GSTR-2B",
    # Excel worksheet titles cannot contain a slash.
    "3 Missing & Invalid GSTIN",
    "4 Matched Entries",
    "5 PR Not in 2B",
    "6 2B Not in PR",
    "7 Date Only",
    "8 PR Reconciled",
    "9 GSTR-2B Reconciled",
]

NAVY = "17324D"
BLUE = "DCEAF7"
PALE_BLUE = "EEF5FB"
GOLD = "F8D66D"
PALE_GOLD = "FFF5CC"
GREEN = "DDEEDB"
WHITE = "FFFFFF"
GRID = "B8C6D1"

THIN_SIDE = Side(style="thin", color=GRID)
THICK_SIDE = Side(style="thick", color=NAVY)
DATE_NUMBER_FORMAT = "dd-mmm-yyyy"
AMOUNT_NUMBER_FORMAT = "#,##0.00;[Red]-#,##0.00"


def write_reconciliation_workbook(
    report: ReconciliationReport,
    output_path: str | Path,
) -> Path:
    """Write all raw, classified, and reconciled views to one XLSX file."""
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "GST Reconciliation Report"
    workbook.properties.subject = "Purchase Register and GSTR-2B reconciliation"

    _write_simple_sheet(
        workbook.create_sheet(SHEET_NAMES[0]),
        report.pr_raw,
    )
    _write_simple_sheet(
        workbook.create_sheet(SHEET_NAMES[1]),
        report.gstr2b_raw,
    )

    _write_simple_sheet(
        workbook.create_sheet(SHEET_NAMES[2]),
        _gstin_review_dataframe(report),
    )

    _write_grouped_sheet(
        workbook.create_sheet(SHEET_NAMES[3]),
        left_label="Purchase Register",
        left_index_label="PR Index",
        left_frame=report.pr_raw,
        center_label="Match Score",
        right_label="GSTR-2B",
        right_index_label="2B Index",
        right_frame=report.gstr2b_raw,
        rows=[(pair.pr_index, pair.score, pair.gstr2b_index) for pair in report.matched_pairs],
        emphasize_center=True,
    )

    _write_grouped_sheet(
        workbook.create_sheet(SHEET_NAMES[4]),
        left_label="Purchase Register - Not in 2B",
        left_index_label="PR Index",
        left_frame=report.pr_raw,
        center_label="Probable Score",
        right_label="Probable GSTR-2B Match",
        right_index_label="2B Index",
        right_frame=report.gstr2b_raw,
        rows=[
            _probable_row(index, report.pr_probable_matches.get(index), reverse=False)
            for index in report.pr_unmatched_indexes
        ],
        emphasize_center=True,
    )

    _write_grouped_sheet(
        workbook.create_sheet(SHEET_NAMES[5]),
        left_label="GSTR-2B - Not in Purchase Register",
        left_index_label="2B Index",
        left_frame=report.gstr2b_raw,
        center_label="Probable Score",
        right_label="Probable Purchase Register Match",
        right_index_label="PR Index",
        right_frame=report.pr_raw,
        rows=[
            _probable_row(index, report.gstr2b_probable_matches.get(index), reverse=True)
            for index in report.gstr2b_unmatched_indexes
        ],
        emphasize_center=True,
    )

    _write_grouped_sheet(
        workbook.create_sheet(SHEET_NAMES[6]),
        left_label="Purchase Register",
        left_index_label="PR Index",
        left_frame=report.pr_raw,
        center_label="Date Difference (Days)",
        right_label="GSTR-2B",
        right_index_label="2B Index",
        right_frame=report.gstr2b_raw,
        rows=[
            (
                pair.pr_index,
                _date_difference_days(report, pair),
                pair.gstr2b_index,
            )
            for pair in report.date_only_pairs
        ],
        emphasize_center=False,
    )

    _write_simple_sheet(
        workbook.create_sheet(SHEET_NAMES[7]),
        report.pr_result,
    )
    _write_simple_sheet(
        workbook.create_sheet(SHEET_NAMES[8]),
        report.gstr2b_result,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _gstin_review_dataframe(report: ReconciliationReport) -> pd.DataFrame:
    metadata_columns = ["Source", "Original Index", "GSTIN Issue", "Original GSTIN"]
    source_columns = _review_source_columns(report.pr_raw, metadata_columns)
    source_columns.extend(
        column
        for column in _review_source_columns(report.gstr2b_raw, metadata_columns)
        if column not in source_columns
    )
    records: list[dict] = []

    def add_records(
        source: str,
        raw_frame: pd.DataFrame,
        issues: dict[int, str],
        original_gstin_values: dict[int, object],
    ) -> None:
        for index, issue in sorted(issues.items()):
            record = {
                "Source": source,
                "Original Index": index,
                "GSTIN Issue": issue,
                "Original GSTIN": original_gstin_values.get(index),
            }
            for column, value in raw_frame.loc[index].items():
                record[_review_column_name(column, metadata_columns)] = value
            records.append(record)

    add_records(
        "Purchase Register",
        report.pr_raw,
        report.pr_gstin_issues,
        report.pr_original_gstin_values,
    )
    add_records(
        "GSTR-2B",
        report.gstr2b_raw,
        report.gstr2b_gstin_issues,
        report.gstr2b_original_gstin_values,
    )
    return pd.DataFrame(records, columns=[*metadata_columns, *source_columns])


def _review_source_columns(
    dataframe: pd.DataFrame,
    metadata_columns: list[str],
) -> list[str]:
    return [
        _review_column_name(column, metadata_columns)
        for column in dataframe.columns
    ]


def _review_column_name(column: object, metadata_columns: list[str]) -> str:
    name = str(column)
    return f"Raw {name}" if name in metadata_columns else name


def _probable_row(
    unmatched_index: int,
    probable: MatchPair | None,
    *,
    reverse: bool,
) -> tuple[int, float | None, int | None]:
    if probable is None:
        return unmatched_index, None, None
    probable_index = probable.pr_index if reverse else probable.gstr2b_index
    return unmatched_index, probable.score, probable_index


def _date_difference_days(report: ReconciliationReport, pair: MatchPair) -> int | None:
    pr_value = report.pr_result.at[pair.pr_index, "Invoice Date"]
    twob_value = report.gstr2b_result.at[pair.gstr2b_index, "Invoice Date"]
    try:
        pr_date = pd.to_datetime(str(int(float(pr_value))), format="%Y%m%d")
        twob_date = pd.to_datetime(str(int(float(twob_value))), format="%Y%m%d")
    except (TypeError, ValueError, OverflowError):
        return None
    return abs((twob_date - pr_date).days)


def _write_simple_sheet(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    headers = [str(column) for column in dataframe.columns]
    for column_number, header in enumerate(headers, start=1):
        cell = _set_cell(worksheet, 1, column_number, header)
        _style_header(cell)

    for row_number, row_values in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_number, value in enumerate(row_values, start=1):
            _set_cell(
                worksheet,
                row_number,
                column_number,
                value,
                header=headers[column_number - 1],
            )

    if headers:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{max(1, worksheet.max_row)}"
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 24
    _set_column_widths(worksheet)


def _write_grouped_sheet(
    worksheet: Worksheet,
    *,
    left_label: str,
    left_index_label: str,
    left_frame: pd.DataFrame,
    center_label: str,
    right_label: str,
    right_index_label: str,
    right_frame: pd.DataFrame,
    rows: Iterable[tuple[int, float | int | None, int | None]],
    emphasize_center: bool,
) -> None:
    left_headers = [left_index_label, *[str(column) for column in left_frame.columns]]
    right_headers = [right_index_label, *[str(column) for column in right_frame.columns]]
    left_start = 1
    left_end = len(left_headers)
    center_column = left_end + 1
    right_start = center_column + 1
    right_end = right_start + len(right_headers) - 1

    _merge_and_label(worksheet, 1, left_start, left_end, left_label, BLUE)
    _merge_and_label(worksheet, 1, center_column, center_column, "Match", GOLD)
    _merge_and_label(worksheet, 1, right_start, right_end, right_label, GREEN)

    all_headers = [*left_headers, center_label, *right_headers]
    for column_number, header in enumerate(all_headers, start=1):
        cell = _set_cell(worksheet, 2, column_number, header)
        _style_header(cell)

    for row_number, (left_index, center_value, right_index) in enumerate(rows, start=3):
        left_values = [left_index, *_frame_row_values(left_frame, left_index)]
        right_values = (
            [right_index, *_frame_row_values(right_frame, right_index)]
            if right_index is not None
            else [None] * len(right_headers)
        )
        values = [*left_values, center_value, *right_values]
        for column_number, value in enumerate(values, start=1):
            cell = _set_cell(
                worksheet,
                row_number,
                column_number,
                value,
                header=all_headers[column_number - 1],
            )
            if column_number == center_column and center_value is not None:
                cell.number_format = "0.00" if emphasize_center else "0"

    worksheet.auto_filter.ref = (
        f"A2:{get_column_letter(right_end)}{max(2, worksheet.max_row)}"
    )
    worksheet.freeze_panes = "A3"
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 24

    if emphasize_center:
        _emphasize_score_column(worksheet, center_column)
    else:
        for row_number in range(1, max(2, worksheet.max_row) + 1):
            cell = worksheet.cell(row=row_number, column=center_column)
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
            cell.font = Font(bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _set_column_widths(worksheet)


def _frame_row_values(dataframe: pd.DataFrame, index: int | None) -> list:
    if index is None:
        return [None] * len(dataframe.columns)
    return dataframe.loc[index].tolist()


def _merge_and_label(
    worksheet: Worksheet,
    row: int,
    start_column: int,
    end_column: int,
    label: str,
    color: str,
) -> None:
    if end_column > start_column:
        worksheet.merge_cells(
            start_row=row,
            start_column=start_column,
            end_row=row,
            end_column=end_column,
        )
    cell = _set_cell(worksheet, row, start_column, label)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=NAVY, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_number in range(start_column, end_column + 1):
        worksheet.cell(row=row, column=column_number).fill = PatternFill(
            "solid", fgColor=color
        )


def _style_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=THIN_SIDE)


def _emphasize_score_column(worksheet: Worksheet, column_number: int) -> None:
    for row_number in range(1, max(2, worksheet.max_row) + 1):
        cell = worksheet.cell(row=row_number, column=column_number)
        cell.fill = PatternFill("solid", fgColor=GOLD if row_number <= 2 else PALE_GOLD)
        cell.font = Font(bold=True, color=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=THICK_SIDE, right=THICK_SIDE)


def _set_column_widths(worksheet: Worksheet) -> None:
    for column_number in range(1, worksheet.max_column + 1):
        maximum = 0
        for row_number in range(1, min(worksheet.max_row, 250) + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            if cell.coordinate in worksheet.merged_cells:
                continue
            value = cell.value
            if value is not None:
                maximum = max(maximum, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(maximum + 2, 12), 40
        )


def _set_cell(
    worksheet: Worksheet,
    row: int,
    column: int,
    value: object,
    *,
    header: str | None = None,
):
    clean_value = _display_value(value, header)
    cell = worksheet.cell(row=row, column=column, value=clean_value)
    # CSV text beginning with '=' must remain literal upload data, not become
    # an executable Excel formula in the generated report.
    if isinstance(clean_value, str) and clean_value.startswith("="):
        cell.data_type = "s"
    if _is_date_header(header) and isinstance(clean_value, (datetime, date)):
        cell.number_format = DATE_NUMBER_FORMAT
    elif _is_amount_header(header) and isinstance(clean_value, (int, float)):
        cell.number_format = AMOUNT_NUMBER_FORMAT
    return cell


def _display_value(value: object, header: str | None):
    clean_value = _excel_value(value)
    if _is_date_header(header):
        parsed_date = _workbook_date_value(clean_value)
        if parsed_date is not None:
            return parsed_date.to_pydatetime()
    if _is_amount_header(header):
        parsed_amount = _workbook_amount_value(clean_value)
        if parsed_amount is not None:
            return parsed_amount
    return clean_value


def _is_date_header(header: str | None) -> bool:
    if not header:
        return False
    normalized = " ".join(header.lower().split())
    return "date" in normalized and "difference" not in normalized


def _is_amount_header(header: str | None) -> bool:
    if not header:
        return False
    normalized = " ".join(header.lower().split())
    return normalized in {"taxable value", "igst", "cgst", "sgst", "total gst"} or "taxable" in normalized


def _workbook_date_value(value: object) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value).normalize()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _workbook_numeric_date(float(value))

    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return _workbook_numeric_date(float(text))
    for format_string in (
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d %b %Y", "%d %B %Y",
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
    ):
        parsed = pd.to_datetime(text, format=format_string, errors="coerce")
        if not pd.isna(parsed):
            return parsed.normalize()
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return None if pd.isna(parsed) else parsed.normalize()


def _workbook_numeric_date(value: float) -> pd.Timestamp | None:
    if np.isnan(value):
        return None
    whole_value = int(value)
    if value.is_integer() and 19000101 <= whole_value <= 29991231:
        parsed = pd.to_datetime(str(whole_value), format="%Y%m%d", errors="coerce")
    elif 1 <= value <= 100000:
        parsed = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
    else:
        return None
    return None if pd.isna(parsed) else parsed.normalize()


def _workbook_amount_value(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return None if isinstance(value, float) and np.isnan(value) else float(value)
    text = str(value).strip().replace("\u00a0", "")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = re.sub(r"[,$₹\s]", "", text)
    if text.endswith("-"):
        text = f"-{text[:-1]}"
    try:
        amount = float(text)
    except (TypeError, ValueError):
        return None
    return -amount if negative else amount


def _excel_value(value: object):
    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (datetime, date, str, int, float, bool)):
        if isinstance(value, float) and pd.isna(value):
            return None
        return value
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    return str(value)
