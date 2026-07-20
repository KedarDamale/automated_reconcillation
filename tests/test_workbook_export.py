import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from reconciliation import INVALID_GSTIN, MISSING_GSTIN, run_reconciliation_report
from workbook_export import SHEET_NAMES, write_reconciliation_workbook


HEADER = (
    "Supplier Name,GSTIN,Invoice No,Invoice Date,HSN Code (optional),"
    "Taxable Value,IGST,CGST,SGST,Total GST\n"
)


def row(supplier, gstin, invoice, invoice_date, hsn="0012", taxable=1000,
        igst=0, cgst=90, sgst=90, total=180):
    return (
        f"{supplier},{gstin},{invoice},{invoice_date},{hsn},"
        f"{taxable},{igst},{cgst},{sgst},{total}"
    )


class NineSheetWorkbookTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.directory = Path(self.temp.name)

    def tearDown(self):
        self.temp.cleanup()

    def _write_input(self, name, rows):
        path = self.directory / name
        path.write_text(HEADER + "\n".join(rows), encoding="utf-8")
        return path

    def _build_report(self, *, date_tolerance_days=10):
        purchase_register = self._write_input(
            "purchase.csv",
            [
                row("ACME Exact", "22AAAAA0000A1Z5", "000123", "01/06/2026"),
                row("Date Only", "33BBBBB1111B2Z6", "DATE-1", "01/06/2026"),
                row("Purchase Weak", "44CCCCC2222C3Z7", "PR-001", "03/06/2026"),
                row("No GSTIN", " ", "NULL-1", "04/06/2026"),
                row("PR Only", "55DDDDD3333D4Z8", "ONLY-PR", "05/06/2026"),
            ],
        )
        gstr2b = self._write_input(
            "gstr2b.csv",
            [
                row("ACME Exact", "22AAAAA0000A1Z5", "000123", "01/06/2026"),
                row("Date Only", "33BBBBB1111B2Z6", "DATE-1", "06/06/2026"),
                row("Unrelated Vendor", "44CCCCC2222C3Z7", "OTHER-999", "03/06/2026"),
                row("2B Only", "66EEEEE4444E5Z9", "ONLY-2B", "07/06/2026"),
            ],
        )
        return run_reconciliation_report(
            purchase_register,
            gstr2b,
            date_tolerance_days=date_tolerance_days,
        )

    def test_report_categories_are_mutually_exclusive(self):
        report = self._build_report()

        self.assertEqual([(pair.pr_index, pair.gstr2b_index) for pair in report.matched_pairs], [(0, 0)])
        self.assertEqual([(pair.pr_index, pair.gstr2b_index) for pair in report.date_only_pairs], [(1, 1)])
        self.assertEqual(report.pr_missing_gstin_indexes, [3])
        self.assertEqual(report.pr_unmatched_indexes, [2, 4])
        self.assertEqual(report.gstr2b_unmatched_indexes, [2, 3])
        self.assertEqual(report.pr_probable_matches[2].gstr2b_index, 2)
        self.assertLess(report.pr_probable_matches[2].score, 80)
        self.assertNotIn(4, report.pr_probable_matches)
        self.assertEqual(report.gstr2b_probable_matches[2].pr_index, 2)
        self.assertNotIn(3, report.gstr2b_probable_matches)

    def test_raw_values_and_workbook_layout_are_preserved(self):
        report = self._build_report()
        self.assertEqual(report.pr_raw.loc[0, "Supplier Name"], "ACME Exact")
        self.assertEqual(report.pr_raw.loc[0, "Invoice No"], "000123")
        self.assertEqual(report.pr_raw.loc[0, "HSN Code (optional)"], "0012")

        output = write_reconciliation_workbook(report, self.directory / "report.xlsx")
        workbook = load_workbook(output)
        self.assertEqual(workbook.sheetnames, SHEET_NAMES)

        original = workbook["1 Purchase Register"]
        self.assertEqual(original["A2"].value, "ACME Exact")
        self.assertEqual(original["C2"].value, "000123")
        self.assertEqual(original["E2"].value, "0012")
        self.assertEqual(original["D2"].value, datetime(2026, 6, 1))
        self.assertEqual(original["D2"].number_format, "dd-mmm-yyyy")
        self.assertEqual(original["F2"].value, 1000)
        self.assertEqual(original["F2"].number_format, "#,##0.00;[Red]-#,##0.00")

        missing = workbook["3 Missing & Invalid GSTIN"]
        self.assertEqual(missing["A2"].value, "Purchase Register")
        self.assertEqual(missing["B2"].value, 3)
        self.assertEqual(missing["C2"].value, MISSING_GSTIN)
        self.assertEqual(missing["D2"].value, " ")
        self.assertEqual(missing["E2"].value, "No GSTIN")

        matched = workbook["4 Matched Entries"]
        self.assertEqual(matched["A1"].value, "Purchase Register")
        self.assertEqual(matched["L2"].value, "Match Score")
        self.assertEqual(matched["A3"].value, 0)
        self.assertEqual(matched["L3"].value, 100)
        self.assertEqual(matched["M3"].value, 0)
        self.assertTrue(matched["L3"].font.bold)
        self.assertEqual(matched["L3"].border.left.style, "thick")
        self.assertEqual(matched["L3"].border.right.style, "thick")

        pr_unmatched = workbook["5 PR Not in 2B"]
        self.assertEqual(pr_unmatched["A3"].value, 2)
        self.assertLess(pr_unmatched["L3"].value, 80)
        self.assertEqual(pr_unmatched["M3"].value, 2)
        self.assertEqual(pr_unmatched["A4"].value, 4)
        self.assertIsNone(pr_unmatched["L4"].value)
        self.assertIsNone(pr_unmatched["M4"].value)

        date_only = workbook["7 Date Only"]
        self.assertEqual(date_only["A3"].value, 1)
        self.assertEqual(date_only["L2"].value, "Date Difference (Days)")
        self.assertEqual(date_only["L3"].value, 5)
        self.assertEqual(date_only["M3"].value, 1)

        pr_reconciled = workbook["8 PR Reconciled"]
        self.assertEqual(pr_reconciled["A2"].value, "acme exact")
        pr_headers = [cell.value for cell in pr_reconciled[1]]
        self.assertIn("Best score", pr_headers)
        self.assertIn("Best match 2B index", pr_headers)
        self.assertIn("Match category", pr_headers)
        self.assertEqual(
            pr_reconciled.cell(2, pr_headers.index("Match category") + 1).value,
            "Matched",
        )
        self.assertEqual(pr_reconciled["D2"].value, datetime(2026, 6, 1))
        self.assertEqual(pr_reconciled["D2"].number_format, "dd-mmm-yyyy")

        gstr2b_reconciled = workbook["9 GSTR-2B Reconciled"]
        self.assertEqual(gstr2b_reconciled["A2"].value, "acme exact")
        gstr2b_headers = [cell.value for cell in gstr2b_reconciled[1]]
        self.assertIn("Best match PR index", gstr2b_headers)
        self.assertIn("Probable PR indexes", gstr2b_headers)
        self.assertIn("Match category", gstr2b_headers)

    def test_noisy_and_invalid_gstins_are_classified_once_for_review(self):
        purchase_register = self._write_input(
            "gstin-pr.csv",
            [
                row("Noisy GSTIN", "27ADZFS0848J1Z8_X000D_", "NOISY-1", "01/06/2026"),
                row("Blank GSTIN", " ", "BLANK-1", "02/06/2026"),
                row("Malformed GSTIN", "GST: NOT-VALID", "BAD-1", "03/06/2026"),
            ],
        )
        gstr2b = self._write_input(
            "gstin-2b.csv",
            [
                row("Noisy GSTIN", "27ADZFS0848J1Z8", "NOISY-1", "01/06/2026"),
                row("Invalid 2B", "27ADZFS0848J1Z8 / 22AAAAA0000A1Z5", "BAD-2", "04/06/2026"),
            ],
        )

        report = run_reconciliation_report(purchase_register, gstr2b)

        self.assertEqual(report.pr_result.at[0, "GSTIN"], "27ADZFS0848J1Z8")
        self.assertEqual([(pair.pr_index, pair.gstr2b_index) for pair in report.matched_pairs], [(0, 0)])
        self.assertEqual(report.pr_gstin_issues, {1: MISSING_GSTIN, 2: INVALID_GSTIN})
        self.assertEqual(report.gstr2b_gstin_issues, {1: INVALID_GSTIN})
        self.assertEqual(report.pr_unmatched_indexes, [])
        self.assertEqual(report.gstr2b_unmatched_indexes, [])

        output = write_reconciliation_workbook(report, self.directory / "gstin-review.xlsx")
        workbook = load_workbook(output)
        review = workbook["3 Missing & Invalid GSTIN"]
        rows = list(review.iter_rows(min_row=2, values_only=True))
        self.assertEqual(
            [(row_values[0], row_values[1], row_values[2]) for row_values in rows],
            [
                ("Purchase Register", 1, MISSING_GSTIN),
                ("Purchase Register", 2, INVALID_GSTIN),
                ("GSTR-2B", 1, INVALID_GSTIN),
            ],
        )
        self.assertEqual(rows[0][3], " ")
        self.assertEqual(rows[1][3], "GST: NOT-VALID")
        self.assertEqual(rows[2][3], "27ADZFS0848J1Z8 / 22AAAAA0000A1Z5")
        self.assertEqual(workbook["5 PR Not in 2B"].max_row, 2)
        self.assertEqual(workbook["6 2B Not in PR"].max_row, 2)

    def test_date_only_requires_all_other_fields_and_obeys_tolerance(self):
        report = self._build_report(date_tolerance_days=4)
        self.assertEqual(report.date_only_pairs, [])
        self.assertIn(1, report.pr_unmatched_indexes)
        self.assertIn(1, report.gstr2b_unmatched_indexes)

        pr = self._write_input(
            "tax-mismatch-pr.csv",
            [row("Tax Difference", "77FFFFF5555F6Z1", "TAX-1", "01/06/2026", cgst=90)],
        )
        twob = self._write_input(
            "tax-mismatch-2b.csv",
            [row("Tax Difference", "77FFFFF5555F6Z1", "TAX-1", "03/06/2026", cgst=91)],
        )
        tax_mismatch = run_reconciliation_report(pr, twob, date_tolerance_days=10)
        self.assertEqual(tax_mismatch.date_only_pairs, [])
        self.assertEqual(len(tax_mismatch.matched_pairs), 1)


if __name__ == "__main__":
    unittest.main()
