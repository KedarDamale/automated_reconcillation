import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app import create_app
from workbook_export import SHEET_NAMES


COLUMNS = (
    "Supplier Name,GSTIN,Invoice No,Invoice Date,HSN Code (optional),"
    "Taxable Value,IGST,CGST,SGST,Total GST\n"
)


def csv_upload(prefix: str, count: int = 12) -> io.BytesIO:
    rows = [
        f"Acme Supplier,22AAAAA0000A1Z5,{prefix}-{number},01/06/2026,1001,"
        f"{1000 + number},0,90,90,180"
        for number in range(count)
    ]
    return io.BytesIO((COLUMNS + "\n".join(rows)).encode("utf-8"))


class ReconciliationAppTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.app = create_app()
        self.app.config.update(TESTING=True, JOBS_DIR=Path(self.temp_dir.name))
        self.client = self.app.test_client()

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_upload_results_pagination_and_exports(self):
        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (csv_upload("INV"), "purchase.csv"),
                "gstr2b_file": (csv_upload("INV"), "gstr2b.csv"),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        result_url = response.headers["Location"]
        page = self.client.get(result_url)
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"12 rows", page.data)
        self.assertIn(b"Page 1 of 2", page.data)

        job_id = result_url.rstrip("/").split("/")[-1]
        export = self.client.get(f"/export/{job_id}")
        self.assertEqual(export.status_code, 200)
        self.assertGreater(len(export.data), 1000)
        workbook = load_workbook(io.BytesIO(export.data))
        self.assertEqual(workbook.sheetnames, SHEET_NAMES)
        matched = workbook["4 Matched Entries"]
        self.assertEqual(matched["L2"].value, "Match Score")
        self.assertEqual(matched["L3"].value, 100)
        self.assertEqual(matched["L3"].border.left.style, "thick")
        self.assertEqual(matched["L3"].border.right.style, "thick")
        export.close()

        self.assertEqual(self.client.get(f"/export/{job_id}/pr").status_code, 404)

        pr_rows = json.loads(
            (Path(self.temp_dir.name) / job_id / "pr_result.json").read_text("utf-8")
        )
        self.assertEqual(len(pr_rows), 12)
        self.assertEqual(pr_rows[0]["Best score"], 100.0)
        self.assertEqual(pr_rows[0]["Best match 2B index"], 0)

    def test_rejects_wrong_extension(self):
        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (io.BytesIO(b"bad"), "purchase.txt"),
                "gstr2b_file": (csv_upload("INV", 1), "gstr2b.csv"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn(b"must be a .csv or .xlsx file", response.data)

    def test_columns_preview_returns_fuzzy_matched_suggestions(self):
        response = self.client.post(
            "/columns",
            data={"file": (csv_upload("INV", 1), "purchase.csv")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertIn("GSTIN", payload["columns"])
        self.assertEqual(payload["suggestions"]["Supplier Name"]["source"], "Supplier Name")
        self.assertGreaterEqual(payload["suggestions"]["Supplier Name"]["score"], 95)
        self.assertIn("Taxable Value", payload["required_columns"])

    def test_columns_preview_rejects_wrong_extension(self):
        response = self.client.post(
            "/columns",
            data={"file": (io.BytesIO(b"bad"), "purchase.txt")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("must be a .csv or .xlsx file", response.get_json()["error"])

    def test_reconcile_honors_a_user_confirmed_column_mapping(self):
        pr_csv = io.BytesIO(
            b"supplier,gst_no,inv_no,inv_date,hsn,taxable,igst,cgst,sgst,total\n"
            b"Acme Supplier,22AAAAA0000A1Z5,INV-1,01/06/2026,1001,1000,0,90,90,180\n"
        )
        pr_mapping = json.dumps(
            {
                "Supplier Name": "supplier",
                "GSTIN": "gst_no",
                "Invoice No": "inv_no",
                "Invoice Date": "inv_date",
                "Taxable Value": "taxable",
            }
        )

        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (pr_csv, "purchase.csv"),
                "gstr2b_file": (csv_upload("INV", 1), "gstr2b.csv"),
                "pr_mapping": pr_mapping,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        job_id = response.headers["Location"].rstrip("/").split("/")[-1]
        pr_rows = json.loads(
            (Path(self.temp_dir.name) / job_id / "pr_result.json").read_text("utf-8")
        )
        # These headers ("supplier", "gst_no", ...) don't fuzzy-match the
        # preferred schema, so a populated "Supplier Name" proves the
        # confirmed mapping -- not automatic matching -- was used.
        self.assertEqual(pr_rows[0]["Supplier Name"], "acme supplier")

    def test_reconcile_sums_multiple_taxable_value_columns_from_mapping(self):
        pr_csv = io.BytesIO(
            b"supplier,gst_no,inv_no,inv_date,taxable_base,taxable_adjustment\n"
            b"Acme Supplier,22AAAAA0000A1Z5,INV-1,01/06/2026,900,100\n"
        )
        pr_mapping = json.dumps(
            {
                "Supplier Name": "supplier",
                "GSTIN": "gst_no",
                "Invoice No": "inv_no",
                "Invoice Date": "inv_date",
                "Taxable Value": ["taxable_base", "taxable_adjustment"],
            }
        )

        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (pr_csv, "purchase.csv"),
                "gstr2b_file": (csv_upload("INV", 1), "gstr2b.csv"),
                "pr_mapping": pr_mapping,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        job_id = response.headers["Location"].rstrip("/").split("/")[-1]
        pr_rows = json.loads(
            (Path(self.temp_dir.name) / job_id / "pr_result.json").read_text("utf-8")
        )
        self.assertEqual(pr_rows[0]["Taxable Value"], 1000)
        self.assertEqual(pr_rows[0]["Best match 2B index"], 0)

    def test_reconcile_rejects_a_mapping_pointing_at_a_missing_column(self):
        pr_mapping = json.dumps(
            {
                "Supplier Name": "does-not-exist",
                "GSTIN": "GSTIN",
                "Invoice No": "Invoice No",
                "Invoice Date": "Invoice Date",
                "Taxable Value": "Taxable Value",
            }
        )

        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (csv_upload("INV", 1), "purchase.csv"),
                "gstr2b_file": (csv_upload("INV", 1), "gstr2b.csv"),
                "pr_mapping": pr_mapping,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn(b"does-not-exist", response.data)

    def test_reconcile_rejects_a_mapping_that_reuses_a_source_column(self):
        pr_mapping = json.dumps(
            {
                "Supplier Name": "GSTIN",
                "GSTIN": "GSTIN",
                "Invoice No": "Invoice No",
                "Invoice Date": "Invoice Date",
                "Taxable Value": "Taxable Value",
            }
        )

        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (csv_upload("INV", 1), "purchase.csv"),
                "gstr2b_file": (csv_upload("INV", 1), "gstr2b.csv"),
                "pr_mapping": pr_mapping,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn(b"mapped once", response.data)

    def test_reconcile_respects_custom_date_tolerance_days(self):
        pr = io.BytesIO(
            (COLUMNS + "Acme Supplier,22AAAAA0000A1Z5,INV-1,01/06/2026,1001,1000,0,90,90,180").encode()
        )
        gstr2b = io.BytesIO(
            (COLUMNS + "Acme Supplier,22AAAAA0000A1Z5,INV-1,12/06/2026,1001,1000,0,90,90,180").encode()
        )

        response = self.client.post(
            "/reconcile",
            data={
                "pr_file": (pr, "purchase.csv"),
                "gstr2b_file": (gstr2b, "gstr2b.csv"),
                "date_tolerance_days": "15",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        job_id = response.headers["Location"].rstrip("/").split("/")[-1]
        pr_rows = json.loads(
            (Path(self.temp_dir.name) / job_id / "pr_result.json").read_text("utf-8")
        )
        # 11 days apart: default 10-day tolerance would leave this unmatched.
        self.assertEqual(pr_rows[0]["Best match 2B index"], 0)


if __name__ == "__main__":
    unittest.main()
